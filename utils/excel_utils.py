import os
import openpyxl
from openpyxl.styles import *
from openpyxl.utils import get_column_letter

class Spreadsheet():
    def __init__(self, filename=None):
        self.filename = filename
        if filename == None:
            print('Creando archivo')
            self.workbook = openpyxl.Workbook()
        else:
            print('Cargando archivo')
            self.workbook = openpyxl.load_workbook(filename)
        self.get_workbook()

    def get_workbook(self):
        return self.workbook

    def get_active_sheet(self):
        self.sheet = self.workbook.active
        return self.workbook.active

    def get_column_letter(self, number):
        return get_column_letter(number)

    def create_sheet(self, name):
        self.workbook.create_sheet(name)
    
    def set_active_sheet(self, name):
        self.sheet = self.workbook[name]

    def set_title(self, title):
        self.get_active_sheet().title = title

    def format(self, cell_range=None):
        first_column = 'A'
        last_column = get_column_letter(self.sheet.max_column)
        if cell_range == None:
            cell_range = f'{first_column}:{last_column}'
    

    def get_filename(self):
        return self.filename

    def set_filename(self, filename):
        self.filename = filename

    def save(self, path="./"):
        print(f'Saving to {path + self.filename}')
        self.workbook.save(filename = path + self.filename)

    def open_spreadsheet(self):
        print('Opening')
        os.startfile(self.filename)

    def set_font(self, cell_range = None, font='Arial', font_size=12, isBold : bool = True, color='0000000', row_count=99):
        font = Font(name=font,
                 size=font_size,
                 bold= isBold,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color=color)
        print(len(self.sheet))
        first_column = 'A'
        last_column = get_column_letter(self.sheet.max_column)
        if cell_range == None:
            cell_range = f'{first_column}1:{last_column}{row_count + 1}'

        for row in self.sheet[cell_range]:
            for cell in row:
                cell.font = font
        print(font.__class__.__base__)
        return font

    def set_borders(self, cell_range=None, border_style="thin", color="000000", row_count=99):
        """
        double
        medium
        thin

        """
        first_column = 'A'
        last_column = get_column_letter(self.sheet.max_column)
        if cell_range == None:
            cell_range = f'{first_column}1:{last_column}{row_count+1}'

        border = Border(left=Side(style=border_style, color=color), 
                     right=Side(style=border_style,color=color), 
                     top=Side(style=border_style,color=color), 
                     bottom=Side(style=border_style,color=color))
        
        for row in self.sheet[cell_range]:
            for cell in row:
                cell.border = border
        print(border.__class__.__base__)
        
        return border
    
    def set_alignment(self, cell_range = None, horizontal='left', vertical='center', row_count=99):
        first_column = 'A'
        last_column = get_column_letter(self.sheet.max_column)

        if cell_range == None:
            cell_range = f'{first_column}1:{last_column}{row_count+1}'
        alignment = Alignment(horizontal=horizontal, vertical=vertical)

        for row in self.sheet[cell_range]:
            for cell in row:
                cell.alignment = alignment
        print(alignment.__class__.__base__)
        return alignment


